from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
import PySimpleGUI as sg      
import sys 
import csv
import re
import os
#TODO: Check if changing the date from MM/DD/YYYY to DD/MM/YYYY affected Calender. 
def createGui():
        timeFormat= [[sg.Text('Choose Time Format', justification='left', size=(18, 1))], 
                        [sg.Combo(['1:00 PM','13:00'], size=(10,1), readonly=True)]]
        dateFormat= [[sg.Text('Choose Date Format',  justification='left', size=(18, 1))], 
                        [sg.Combo(['MM/DD/YYYY','DD/MM/YYYY','YYYY-MM-DD'], size= (10,1),readonly=True)]]
        layout= [
                [sg.Text('Select the POW', size=(32,1)), sg.FileBrowse()],
                [sg.Column(timeFormat),sg.Column(dateFormat)],
                [sg.Submit(), sg.Cancel()]
        ]
        window= sg.Window("POW to google Calender", layout)
        event,(number) = window.Read()
        #window.Close()
        outputLoc= number['Browse']
        outputLoc= outputLoc.rsplit('/',1)
        print(outputLoc[0])
        sg.Popup('CSV file located at'+ outputLoc[0])
        timeFormatSelector=0
        dateFormatSelector=0
        if number[0] == '1:00 PM':
                timeFormatSelector= 2
        elif number[0] == '13:00':
                timeFormatSelector=1
        if number[1] == 'MM/DD/YYYY':
                dateFormatSelector=1
        elif number[1]== 'DD/MM/YYYY':
                dateFormatSelector=2
        elif number[1]== 'YYYY-MM-DD':
                dateFormatSelector=3
        return [timeFormatSelector, dateFormatSelector, number['Browse'],outputLoc[0]]


class CalEvent:
    #contains a class for each event 
    def __init__(self, date, time, name, mustloc, execloc, OIC, personnel, uniform):
        self.date = date
        self.time = time
        self.name = name
        self.mustloc = mustloc
        self.execloc = execloc
        self.personnel = personnel
        self.uniform = uniform




def convertDatetoExcel(date):
    offset= 693594
    substring= date.split(" ")
    itime = datetime.strptime(substring[0], '%Y-%m-%d')
    n = itime.toordinal()
    return (n - offset) 


def convertExceltoDate(serialDate,dateFormat):
    #if dateFormat=2 replace - to / in date
    dt = datetime.fromordinal(datetime(1900,1,1).toordinal() + int(serialDate) - 2)
    d= dt.date()
    #dd/mm/yyyy
    if dateFormat == 1:
            d=str(d).replace('-','/')
            splitStr= d.split('/')
            d= splitStr[2]+'/'+splitStr[1] +'/' +splitStr[0]
    #change to mm/dd/yyyy
    if dateFormat == 2:
            d=str(d).replace('-','/')
            splitStr=d.split('/')
            d= splitStr[1]+'/'+splitStr[2]+'/'+splitStr[0]
    #yyyy-mm-dd
    if dateFormat==3:
            d=str(d)
            splitStr=d.split(' ')
            d= splitStr[0]

    return d 

def timeFormatChange(time, timeFormat):
    #if timeFormat=1 keep 24 hour time
    #if timeFormat=2 change to AM PM time
    if timeFormat==2:
            if int(time)>=1300:
                    newTime = int(time)-1200
                    if len(str(newTime))==3:
                            strTime= str(newTime)
                            strTime= strTime[:1] + ':' + strTime[1:] + ' PM'
                            return strTime
                    else:
                           strTime= str(newTime)
                           strTime= strTime[:2] + ':' + strTime[2:] + ' PM' 
                           return strTime
            else:
                    newTime= str(time)
                    newTime= newTime[:2] + ':' + newTime[2:] + ' AM'
                    return newTime
    else:
            newTime= str(time)
            newTime= newTime[:2] + ':' + newTime[2:]
            return newTime

def createNewSheet(eventList, dateFormat, timeFormat,output_dir):
    newWb= Workbook()
    dest_filename = output_dir + '/formated_sheet.xlsx'
    ws1 = newWb.active
    ws1.title= "google calender formated sheet"
    colnames = ["Subject",'Start date','Start time','End Date','End Time','All Day Event','Description','Location','Private']
    ws1.append(colnames)
    for x in range(2,len(eventList)):
        descript= "Personnel: " + str(eventList[x].personnel) + "; UOD: " + str(eventList[x].uniform) + "; Execution Location: " + str(eventList[x].execloc)
        ws1.cell(None,x,1).value = eventList[x].name
        ws1.cell(None,x,2).value = convertExceltoDate(eventList[x].date, dateFormat)
        ws1.cell(None,x,7).value = descript
        ws1.cell(None,x,8).value = eventList[x].mustloc
        print(ws1.cell(None,x,2).value)
        if not eventList[x].time:
                ws1.cell(None,x,6).value = True
        else:
                print(timeFormatChange(eventList[x].time, timeFormat))
                ws1.cell(None,x,3).value = timeFormatChange(eventList[x].time, timeFormat)



    newWb.save(filename = dest_filename)
    return dest_filename





parameters= createGui()
wb= load_workbook(filename= parameters[2])
sheet = wb.worksheets[0]
row_count = sheet.max_row
#print(str(sheet.cell(None, 15,1).value))
index=[]
events=[]
dates=[]
#pattern used to match date if not in serial date format
pattern = "^((31(?! (Feb|Apr|Jun|Sep|Nov)))|((30|29)(?! Feb))|(29(?= Feb (((1[6-9]|[2-9]\d)(0[48]|[2468][048]|[13579][26])|((16|[2468][048]|[3579][26])00)))))|(0?[1-9])|1\d|2[0-8])-(Jan|Feb|Mar|May|Apr|Jul|Jun|Aug|Oct|Sep|Nov|Dec)-((1[6-9]|[2-9]\d))$"
pattern2= "[1-2][0-1][0-9][0-9]-"

#checks if the column is a date. Assumes that the date is in excell serial format
# puts the index of the date into the list. This shows where to start getting the dates from document

for x in range(1,row_count):
    j= (sheet.cell(None,x,1).value)
    if not (j is None) and (isinstance(j, float) or re.search(pattern2,str(j))) :
        index.append(x)
        dates.append(j)
        print (j)
for x in range(len(index)):
    i=index[x]
    while sheet.cell(None,i,3).value:
        #specDate=dates[x]
        #if date is in format dd-Mon-yy. convert to serial date format for consistency. 
        if re.match(pattern2,str(sheet.cell(None,index[x],1).value)):
                strings= str(sheet.cell(None,index[x],1).value)
                #subs= strings.split("-")
                specDate= convertDatetoExcel(strings)
        else :
                specDate= sheet.cell(None,index[x],1).value
        time= sheet.cell(None,i,2).value
        eventName= sheet.cell(None,i,3).value
        mustloc=sheet.cell(None,i,4).value
        execloc= sheet.cell(None,i,5).value
        oic=sheet.cell(None,i,6).value
        personnel= sheet.cell(None,i,7).value
        uniform= sheet.cell(None,i,8).value
        events.append(CalEvent(specDate,time,eventName,mustloc,execloc,oic,personnel,uniform))
        i+=1

    


file_name= createNewSheet(events, 3, 2,parameters[3])
wb= load_workbook(file_name)
sh = wb.get_active_sheet()
with open(parameters[3]+'/formated_sheet.csv', 'w') as f:
        c = csv.writer(f)
        for r in sh.rows:
                c.writerow([cell.value for cell in r])

os.remove(file_name)

